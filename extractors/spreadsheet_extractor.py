"""
Spreadsheet extractor for the knowledge graph-based business consulting system.
Extracts structured data from Excel and CSV files.
"""

import os
import logging
import json
import csv
import pandas as pd
import openpyxl
from datetime import datetime
from extractors.extractor_utils import (
    ensure_output_dir, generate_output_filename, save_extraction_result,
    create_extraction_result_template, detect_entities, COMMON_ENTITY_PATTERNS
)

# Add missing function for Excel processing
def _process_excel(file_path):
    """
    Process an Excel file (.xlsx or .xls)
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        dict: Extracted data
    """
    # Load workbook
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # Get workbook metadata
    metadata = {
        "file_type": "Excel",
        "sheet_names": wb.sheetnames,
        "sheet_count": len(wb.sheetnames),
        "author": wb.properties.creator or "",
        "created": wb.properties.created.isoformat() if wb.properties.created else "",
        "modified": wb.properties.modified.isoformat() if wb.properties.modified else "",
    }
    
    # Process each sheet
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get dimensions
        min_row, min_col = 1, 1
        max_row, max_col = ws.max_row, ws.max_column
        
        # Extract headers (first row)
        headers = []
        for col in range(min_col, max_col + 1):
            cell_value = ws.cell(min_row, col).value
            headers.append(str(cell_value) if cell_value is not None else "")
        
        # Extract data (remaining rows)
        data = []
        for row in range(min_row + 1, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell_value = ws.cell(row, col).value
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            # Skip empty rows
            if any(cell.strip() for cell in row_data):
                data.append(row_data)
        
        sheets.append({
            "name": sheet_name,
            "headers": headers,
            "data": data,
            "row_count": len(data),
            "column_count": len(headers)
        })
    
    return {
        "metadata": metadata,
        "sheets": sheets
    }

# Set up logging
logging.basicConfig(level=logging.INFO, 
                   format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def extract(file_path, output_dir=None):
    """
    Extract content from a spreadsheet file (Excel or CSV)
    
    Args:
        file_path (str): Path to the spreadsheet file
        output_dir (str, optional): Directory to save parsed data
        
    Returns:
        dict: Extracted content with metadata
    """
    logger.info(f"Extracting content from spreadsheet: {file_path}")
    
    # Validate file exists
    if not os.path.isfile(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Check file extension
    file_ext = os.path.splitext(file_path)[1].lower()
    if file_ext not in ['.xlsx', '.xls', '.csv']:
        raise ValueError(f"Unsupported file type: {file_ext}")
    
    # Ensure output directory exists
    output_dir = ensure_output_dir(output_dir)
    
    # Generate output filename
    name_without_ext, output_filename = generate_output_filename(file_path)
    output_path = os.path.join(output_dir, output_filename)
    
    # Create extraction result template
    result = create_extraction_result_template(file_path, "spreadsheet_extractor")
    
    try:
        # Process based on file type
        if file_ext == '.csv':
            extracted_data = _process_csv(file_path)
        else:  # Excel files (.xlsx, .xls)
            extracted_data = _process_excel(file_path)
        
        # Extract entities from data
        all_text = ""
        for sheet in extracted_data.get("sheets", []):
            # Collect text from headers and data
            if sheet.get("headers"):
                all_text += " ".join(str(h) for h in sheet["headers"]) + " "
            
            # Extract text from all cells
            for row in sheet.get("data", []):
                all_text += " ".join(str(cell) for cell in row) + " "
        
        # Detect entities in the combined text
        entities = detect_entities(all_text, COMMON_ENTITY_PATTERNS)
        
        # Combine results
        result["document_metadata"] = extracted_data.get("metadata", {})
        result["content"]["text"] = all_text
        result["content"]["sheets"] = extracted_data.get("sheets", [])
        result["content"]["tables"] = extracted_data.get("sheets", [])  # For consistency
        result["content"]["entities"] = entities
        
        # Save extraction result
        save_extraction_result(result, output_path)
        
        logger.info(f"Extraction completed successfully. Output saved to: {output_path}")
        return result
    
    except Exception as e:
        logger.error(f"Spreadsheet extraction failed: {e}")
        raise

def _process_csv(file_path):
    """
    Process a CSV file
    
    Args:
        file_path: Path to CSV file
        
    Returns:
        dict: Extracted data
    """
    # Determine encoding and dialect
    encoding = 'utf-8'
    
    # Try to read with different encodings if necessary
    try:
        df = pd.read_csv(file_path, encoding=encoding)
    except UnicodeDecodeError:
        try:
            encoding = 'latin-1'
            df = pd.read_csv(file_path, encoding=encoding)
        except Exception as e:
            logger.warning(f"Failed to read CSV with pandas: {e}")
            # Fallback to manual CSV parsing
            with open(file_path, 'r', encoding=encoding, errors='replace') as f:
                reader = csv.reader(f)
                headers = next(reader, [])
                data = list(reader)
                
                return {
                    "metadata": {
                        "file_type": "CSV",
                        "encoding": encoding,
                        "rows": len(data) + 1,  # +1 for header
                        "columns": len(headers),
                    },
                    "sheets": [{
                        "name": os.path.basename(file_path),
                        "headers": headers,
                        "data": data,
                        "row_count": len(data),
                        "column_count": len(headers)
                    }]
                }
    
    # If pandas succeeded, convert to our format
    headers = df.columns.tolist()
    data = df.values.tolist()
    
    return {
        "metadata": {
            "file_type": "CSV",
            "encoding": encoding,
            "rows": len(data) + 1,  # +1 for header
            "columns": len(headers),
        },
        "sheets": [{
            "name": os.path.basename(file_path),
            "headers": headers,
            "data": data,
            "row_count": len(data),
            "column_count": len(headers)
        }]
    }